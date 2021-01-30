#!/usr/bin/env python
# usage: quality_checks_v2.py
__author__ = "Heiko Maerz"
__copyright__ = "Copyright (c) 2019-2020 Health Data Research UK. All Rights Reserved."
__email__ = "heiko@metadataworks.co.uk"
__license__ = "Apache 2"

# load packages
import copy
import datetime
import json
import jsonschema
import os
import pandas as pd
import requests
import platform
from openpyxl import load_workbook

CWD = os.getcwd()
DM_JSON_PATH = os.path.join(CWD,'datasets.v2.json')
VALIDATION_SCHEMA_PATH = "https://hdruk.github.io/schemata/schema/dataset/latest/dataset.schema.json"
VALIDATION_WEIGHTS_PATH = os.path.join(CWD, 'config', 'weights', 'latest', 'weights.v2.json')
MEDALLIONS = os.path.join(CWD, 'config', 'weights', 'latest', 'medallions.v2.json')


def strip_string_to_alphanum(text_in):
    if not isinstance(text_in, str):
        return text_in

    text_out = ''
    for c in text_in:
        if c.isalnum():
            text_out = f"{text_out}{c.lower()}"
        elif c in (' ', '-', '_', '.', ','):
            text_out = f"{text_out}_"
    text_out = text_out.replace('___', '_')
    text_out = text_out.replace('__', '_')
    return text_out.strip()


def read_publisher(text_in):
    text_out = str(text_in).upper()
    tokens = text_out.split('>')
    if len(tokens) < 2:
        return text_out.strip()
    return tokens[-1].strip()


def strip_breaks(text_in):
    if not isinstance(text_in, str):
        return text_in
    text_in = text_in.replace('"', r'\"')
    text_in = text_in.replace('\t', ' ')
    text_in = ' '.join(text_in.splitlines())
    text_in = text_in.replace(r'\n\"', ' ')
    text_in = text_in.replace(r'\""', ' ')
    # text_in = text_in.replace("'", r'\\u0027')
    text_in = text_in.replace('  ', ' ')
    text_in = text_in.replace('"', "'")
    return text_in.strip()


def remove_none_from_dict(json_data):
    keys = copy.deepcopy(list(json_data.keys()))
    for js_key in keys:
        js_value = json_data.get(js_key, None)
        if isinstance(js_value, dict):
            remove_none_from_dict(js_value)
            if 0 == len(list(js_value.keys())):
                json_data.pop(js_key, None)
        else:
            if not js_value:
                json_data.pop(js_key, None)


def write_header():
    write_timestamp(f"{__file__}")
    print(f"python=={platform.python_version()}")
    print(f"jsonschema=={jsonschema.__version__}")
    print(f"pandas=={pd.__version__}")
    print(f"requests=={requests.__version__}")
    print()


# timestamp
def write_timestamp(out_text=''):
    now = datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    print(f"{now} {out_text}")
    return


def get_json(json_uri):
    if isinstance(json_uri, dict):
        return json_uri
    elif os.path.isfile(json_uri):
        with open(json_uri, 'r') as json_file:
            return json.load(json_file)
    elif json_uri.startswith('http'):
        return requests.get(json_uri).json()
    else:
        raise Exception


def export_json(data, filename, indent=2):
    with open(filename, 'w') as jsonfile:
        json.dump(data, jsonfile, indent=indent)


# write excel
def write_excel(fname, worksheets, idx=False):
    '''
    Write a number of worksheets to an Excel File
    :param fname: file path and file name
    :param worksheets: dictionary: {(worksheet name): (worksheet DataFrame), ...}
    :param idx: Boolean, save DataFrame indes to Excel, default False
    :return: None yet
    '''
    
    with pd.ExcelWriter(fname) as writer:
        for sheetname, df_worksheet in worksheets.items():
            df_worksheet.to_excel(writer, sheet_name=sheetname, index=idx)

    fname = os.path.join(CWD, 'reports', 'metadata_score_breakdown_v2.xlsx')
    wb = load_workbook(filename = fname)
    datasets_sheet = wb['Datasets']
    num_dm = len(datasets_sheet['G']) - 1
    
    links = []
    for i in range(num_dm):
        datasets_sheet.cell(row=i+2, column=7).hyperlink = "#'{}'!A1".format(datasets_sheet['G{}'.format(i+2)].value)

    wb.save(filename = fname)


def get_datamodels(jason_uri):
    raw_data = get_json(jason_uri)
    return raw_data.get('dataModels', [])


def get_validation_weights(val_weights_path):
    raw_weights = get_json(val_weights_path)

    weights = {}
    for _, attributes in raw_weights.items():
        for attr, weight in attributes.items():
            weights[attr] = {'weight': weight, 'score': 0}

    return weights


def flatten_dictionary(i_key, i_dict, c_data):
    for d_key, d_value in i_dict.items():
        c_key = f"{i_key}.{d_key}"
        if isinstance(d_value, dict):
            flatten_dictionary(c_key, d_value, c_data)
        else:
            c_data[c_key] = d_value


def flatten_datamodel(data_model):
    flat_data = {}
    complex_model = copy.deepcopy(data_model)

    metadataCount = complex_model.pop('structuralMetadata', {})
    metadataCount = metadataCount.get('structuralMetadataCount', {})
    smd = {'structuralMetadata.dataClassesCount': metadataCount.get('structuralMetadata.dataClassesCount', 0),
           'structuralMetadata.tableName': metadataCount.get('structuralMetadata.tableName', 0),
           'structuralMetadata.tableDescription': metadataCount.get('structuralMetadata.tableDescription', 0),
           'structuralMetadata.dataElementsCount': metadataCount.get('structuralMetadata.dataElementsCount', 0),
           'structuralMetadata.columnName': metadataCount.get('structuralMetadata.columnName', 0),
           'structuralMetadata.columnDescription': metadataCount.get('structuralMetadata.columnDescription', 0),
           'structuralMetadata.dataType': metadataCount.get('structuralMetadata.dataType', 0),
           'structuralMetadata.sensitive': metadataCount.get('structuralMetadata.sensitive', 0), }

    for dict_key, dict_value in complex_model.items():
        if isinstance(dict_value, dict):
            flatten_dictionary(dict_key, dict_value, flat_data)
        else:
            flat_data[dict_key] = dict_value

    denominator = smd['structuralMetadata.dataClassesCount']
    if denominator > 0:
        flat_data['structuralMetadata.dataClassesCount'] = 1
        flat_data['structuralMetadata.tableName'] = smd['structuralMetadata.tableName'] / denominator
        flat_data['structuralMetadata.tableDescription'] = smd['structuralMetadata.tableDescription'] / denominator
    else:
        flat_data['structuralMetadata.dataClassesCount'] = 0
        flat_data['structuralMetadata.tableName'] = 0
        flat_data['structuralMetadata.tableDescription'] = 0
    denominator = smd['structuralMetadata.dataElementsCount']
    if denominator > 0:
        flat_data['structuralMetadata.columnName'] = smd['structuralMetadata.columnName'] / denominator
        flat_data['structuralMetadata.columnDescription'] = smd['structuralMetadata.columnDescription'] / denominator
        flat_data['structuralMetadata.dataType'] = smd['structuralMetadata.dataType'] / denominator
        flat_data['structuralMetadata.sensitive'] = smd['structuralMetadata.sensitive'] / denominator
    else:
        flat_data['structuralMetadata.columnName'] = 0
        flat_data['structuralMetadata.columnDescription'] = 0
        flat_data['structuralMetadata.dataType'] = 0
        flat_data['structuralMetadata.sensitive'] = 0

    return flat_data


def assess_completeness(completeness, data_model):
    dm_completeness = copy.deepcopy(completeness)
    flat_dm = flatten_datamodel(data_model)

    total_count, total_weight = 0, 0
    for comp_key, comp_score in dm_completeness.items():
        dm_data = flat_dm.get(comp_key, None)
        if 'identifier' == comp_key:
            dm_completeness[comp_key]['value'] = f"{dm_data}"[-36:]
            # if flat_dm.get('HOP_status', None):
            dm_completeness[comp_key]['score'] = comp_score['weight']
            total_count += 1
            total_weight += comp_score['weight']
            continue
        elif 'structuralMetadata' == comp_key[:18]:
            dm_completeness[comp_key]['score'] = dm_data * comp_score['weight']
            dm_completeness[comp_key]['value'] = f"{dm_data}"[:64]
            if dm_data > 0:
                total_count += 1
            total_weight += dm_data * comp_score['weight']
            continue
        if dm_data:
            dm_completeness[comp_key]['score'] = comp_score['weight']
            dm_completeness[comp_key]['value'] = f"{dm_data}"[:64]
            total_count += 1
            total_weight += comp_score['weight']

    # special rules
    # - continuous data collection
    if flat_dm.get('provenance.temporal.accrualPeriodicity', None):
        if 'CONTINUOUS' == flat_dm['provenance.temporal.accrualPeriodicity']:
            dm_completeness['provenance.temporal.endDate']['score'] = dm_completeness['provenance.temporal.endDate'][
                'weight']
            dm_completeness['provenance.temporal.endDate']['value'] = 'continuous data collection'
            total_count += 1
            total_weight += dm_completeness['provenance.temporal.endDate']['score']
            dm_completeness['provenance.temporal.distributionReleaseDate']['score'] = \
            dm_completeness['provenance.temporal.distributionReleaseDate']['weight']
            dm_completeness['provenance.temporal.distributionReleaseDate']['value'] = 'continuous data collection'
            total_count += 1
            total_weight += dm_completeness['provenance.temporal.distributionReleaseDate']['score']

    return {'count': total_count, 'weight': total_weight, 'completeness': dm_completeness}


def assess_errors(validator, validation_errors, data_model):
    dm_errors = copy.deepcopy(validation_errors)

    # special rules
    error_exceptions = ['accessibility.usage.isReferencedBy']
    # - continuous data collection
    if data_model.get('provenance', None):
        if data_model['provenance'].get('temporal', None):
            if 'CONTINUOUS' == data_model['provenance']['temporal'].get('accrualPeriodicity', ''):
                error_exceptions.extend(
                    ['provenance.temporal.accrualPeriodicity', 'provenance.temporal.distributionReleaseDate'])
            if 'IRREGULAR' == data_model['provenance']['temporal'].get('accrualPeriodicity', ''):
                error_exceptions.extend(
                    ['provenance.temporal.accrualPeriodicity', 'provenance.temporal.distributionReleaseDate'])

    error_count, error_weight = 0, 0
    errors = sorted(validator.iter_errors(data_model), key=lambda e: e.path)
    for e in errors:
        validation_path = [f"{token}" for token in list(e.absolute_path)]
        if len(validation_path) < 1:
            error_key = ''
        else:
            error_key = '.'.join(validation_path)
        # if len(e.absolute_path) > 0:
        #     error_key = '.'.join(e.absolute_path)
        error_msg = e.message.split("'")
        if len(error_msg) > 1:
            test_key = error_msg[1]
            if len(error_key) > 0:
                test_key = f"{error_key}.{test_key}"
            if dm_errors.get(test_key, None):
                error_key = test_key
        if dm_errors.get(error_key, None):
            if error_key not in error_exceptions:
                dm_errors[error_key]['score'] = dm_errors[error_key]['weight']
                error_count += 1
                error_weight += dm_errors[error_key]['weight']
                dm_errors[error_key]['err_msg'] = e.message[-128:]

    metadataCount = data_model.get('structuralMetadata', {})
    metadataCount = metadataCount.get('structuralMetadataCount', {})
    smd = {'structuralMetadata.dataClassesCount': metadataCount.get('structuralMetadata.dataClassesCount', 0),
           'structuralMetadata.tableName': metadataCount.get('structuralMetadata.tableName', 0),
           'structuralMetadata.tableDescription': metadataCount.get('structuralMetadata.tableDescription', 0),
           'structuralMetadata.dataElementsCount': metadataCount.get('structuralMetadata.dataElementsCount', 0),
           'structuralMetadata.columnName': metadataCount.get('structuralMetadata.columnName', 0),
           'structuralMetadata.columnDescription': metadataCount.get('structuralMetadata.columnDescription', 0),
           'structuralMetadata.dataType': metadataCount.get('structuralMetadata.dataType', 0),
           'structuralMetadata.sensitive': metadataCount.get('structuralMetadata.sensitive', 0), }

    denominator = smd['structuralMetadata.dataClassesCount']
    if denominator > 0:
        if denominator > smd['structuralMetadata.tableName']:
            score = (denominator - smd['structuralMetadata.tableName']) / denominator
            dm_errors['structuralMetadata.tableName']['score'] = score * dm_errors['structuralMetadata.tableName'][
                'weight']
            error_count += 1
            error_weight += dm_errors['structuralMetadata.tableName']['score']
        if denominator > smd['structuralMetadata.tableDescription']:
            score = (denominator - smd['structuralMetadata.tableDescription']) / denominator
            dm_errors['structuralMetadata.tableDescription']['score'] = score * dm_errors[
                'structuralMetadata.tableDescription']['weight']
            error_count += 1
            error_weight += dm_errors['structuralMetadata.tableDescription']['score']
    else:
        dm_errors['structuralMetadata.dataClassesCount']['score'] = dm_errors['structuralMetadata.dataClassesCount'][
            'weight']
        error_count += 1
        error_weight += dm_errors['structuralMetadata.dataClassesCount']['weight']
        dm_errors['structuralMetadata.tableName']['score'] = dm_errors['structuralMetadata.tableName']['weight']
        error_count += 1
        error_weight += dm_errors['structuralMetadata.tableName']['weight']
        dm_errors['structuralMetadata.tableDescription']['score'] = dm_errors['structuralMetadata.tableDescription'][
            'weight']
        error_count += 1
        error_weight += dm_errors['structuralMetadata.tableDescription']['weight']
    denominator = smd['structuralMetadata.dataElementsCount']
    if denominator > 0:
        if denominator > smd['structuralMetadata.columnName']:
            score = (denominator - smd['structuralMetadata.columnName']) / denominator
            dm_errors['structuralMetadata.columnName']['score'] = score * dm_errors['structuralMetadata.columnName'][
                'weight']
            error_count += 1
            error_weight += dm_errors['structuralMetadata.columnName']['score']
        if denominator > smd['structuralMetadata.columnDescription']:
            score = (denominator - smd['structuralMetadata.columnDescription']) / denominator
            dm_errors['structuralMetadata.columnDescription']['score'] = score * dm_errors[
                'structuralMetadata.columnDescription']['weight']
            error_count += 1
            error_weight += dm_errors['structuralMetadata.columnDescription']['score']
        if denominator > smd['structuralMetadata.dataType']:
            score = (denominator - smd['structuralMetadata.dataType']) / denominator
            dm_errors['structuralMetadata.dataType']['score'] = score * dm_errors['structuralMetadata.dataType'][
                'weight']
            error_count += 1
            error_weight += dm_errors['structuralMetadata.dataType']['score']
    else:
        dm_errors['structuralMetadata.columnName']['score'] = dm_errors['structuralMetadata.columnName']['weight']
        error_count += 1
        error_weight += dm_errors['structuralMetadata.columnName']['score']
        dm_errors['structuralMetadata.columnDescription']['score'] = dm_errors['structuralMetadata.columnDescription'][
            'weight']
        error_count += 1
        error_weight += dm_errors['structuralMetadata.columnDescription']['score']
        dm_errors['structuralMetadata.dataType']['score'] = dm_errors['structuralMetadata.dataType']['weight']
        error_count += 1
        error_weight += dm_errors['structuralMetadata.dataType']['score']

    return {'count': error_count, 'weight': error_weight, 'errors': dm_errors}


def explain_score(dm_completeness, dm_errors):
    score_details = {'Attribute': [],
                     'Weight': [],
                     'Data (abb)': [],
                     'Completeness': [],
                     'Error': [],
                     'Msg (abb)': []}

    aardvark = {}
    for c_key, c_value in dm_completeness['completeness'].items():
        aardvark[c_key] = {'Attribute': c_key,
                           'Weight': c_value['weight'],
                           'Data (abb)': c_value.get('value', ''),
                           'Completeness': c_value['score']}
        error_data = dm_errors['errors'].get(c_key, None)
        if error_data:
            aardvark[c_key]['Error'] = error_data['score']
            aardvark[c_key]['Msg (abb)'] = error_data.get('err_msg', '')

    for _, aard in aardvark.items():
        for d_key, d_value in aard.items():
            score_details[d_key].append(d_value)

    return score_details


def determine_medallion(medallions, dm_score):
    for medallion, rules in medallions.items():
        if rules['min excluding'] < dm_score['weighted_quality_score'] <= rules['max including']:
            dm_score['weighted_quality_rating'] = medallion
            return
    return


def score_data_models(val_schema_path, val_weights_path, m_path, data_models, debug_out=False):
    write_timestamp(f"scoring {len(data_models)} datasets")
    validation_schema = get_json(val_schema_path)
    dm_validator = jsonschema.Draft7Validator(validation_schema, format_checker=jsonschema.draft7_format_checker)
    completeness = get_validation_weights(val_weights_path)
    validation_errors = get_validation_weights(val_weights_path)
    medallions = get_json(m_path)

    score_json = {'schema_version': '2.0.1',
                'pid': '',
                'id': '',
                'publisher': '',
                'title': '',
                'weighted_quality_rating': 'Not Rated',
                'weighted_quality_score': 0,
                'weighted_completeness_percent': 0,
                'weighted_error_percent': 0
               }
    dm_scores = []
    all_scores = {'Organisation': [],
                  'Title': [],
                  'id': [],
                  'Completeness': [],
                  'Errors': [],
                  'Score': [],
                  'ref': []}
    excel_score = {'Datasets': None}
    reference_counter = len(data_models)
    for data_model in data_models:
        if not data_model.get('id', None):
            write_timestamp(
                f"  ERR: no id for {data_model['summary']['publisher']['name']}>'{data_model['summary']['title']}'")
            continue
        # if 'NHS DIGITAL'!=data_model['summary']['publisher']['name'].upper():
        #     continue
        dm_score = copy.deepcopy(score_json)
        dm_score['pid'] = data_model['pid']
        dm_score['id'] = data_model['id']
        dm_score['publisher'] = f"{data_model['summary']['publisher']['memberOf']} > {data_model['summary']['publisher']['name']}"
        dm_score['title'] = data_model['summary']['title']
        dm_completeness = assess_completeness(completeness, data_model)
        dm_errors = assess_errors(dm_validator, validation_errors, data_model)
        all_scores['Organisation'].append(data_model['summary']['publisher'].get('name', 'no org'))
        all_scores['Title'].append(data_model['summary'].get('title', 'no title'))
        all_scores['id'].append(data_model['id'])
        cmpl_sc = (100 * dm_completeness['weight'])
        dm_score['weighted_completeness_percent'] = round(cmpl_sc, 2)
        all_scores['Completeness'].append(f"{cmpl_sc:.2f}%")
        err_sc = (100 * dm_errors['weight'])
        dm_score['weighted_error_percent'] = round(err_sc, 2)
        all_scores['Errors'].append(f"{err_sc:.2f}%")
        total_sc = 50 * ((dm_completeness['weight']) + (1 - dm_errors['weight']))
        dm_score['weighted_quality_score'] = round(total_sc, 2)
        determine_medallion(medallions, dm_score)
        all_scores['Score'].append(f"{total_sc:.2f}%")
        reference_key = f"data-model-{reference_counter:04d}"
        reference_counter -= 1
        all_scores['ref'].append(reference_key)
        write_timestamp(
            f"{reference_counter:04}-{data_model['summary']['publisher'].get('name', 'no org')}>'{data_model['summary'].get('title', 'no title')}': cmp={cmpl_sc:.2f}%, err={err_sc:.2f}%")
        sc_details = explain_score(dm_completeness, dm_errors)
        excel_score[reference_key] = pd.DataFrame(sc_details)
        dm_scores.append(dm_score)

    if debug_out:
        excel_score['Datasets'] = pd.DataFrame(all_scores)
        print()
        fname = os.path.join(CWD, 'reports', 'metadata_score_breakdown_v2.xlsx')
        write_timestamp(f"{fname}")
        write_excel(fname, excel_score)

    return dm_scores


def write_metadata_quality(dm_scores):
    fname = os.path.join(CWD, 'reports', 'latest', 'metadata_quality.v2.json')
    write_timestamp(f"{fname}")
    export_json(dm_scores, fname)

    csv_out = {}
    for dm in dm_scores:
        for out_key, out_value in dm.items():
            if not csv_out.get(out_key, None):
                csv_out[out_key] = []
            csv_out[out_key].append(out_value)
    df_out = pd.DataFrame(csv_out)
    fname = os.path.join(CWD, 'reports', 'latest', 'metadata_quality.v2.csv')
    write_timestamp(f"{fname}")
    df_out.to_csv(fname, index=False)
    print()
    return


def main():
    write_header()

    data_models = get_datamodels(DM_JSON_PATH)

    dm_scores = score_data_models(VALIDATION_SCHEMA_PATH, VALIDATION_WEIGHTS_PATH, MEDALLIONS, data_models, True)

    write_metadata_quality(dm_scores)

    # TODO: Incorporate Utility scores from V1

    write_timestamp(f"done")


if '__main__' == __name__:
    main()
    print(f" bye ...")